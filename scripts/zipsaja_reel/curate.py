#!/usr/bin/env python3
"""
Excel 매물 데이터 → zipsaja 릴스용 JSON 큐레이션.
Usage:
  python3 scripts/zipsaja_reel/curate.py \
      --excel /path/to/매물.xlsx \
      --gu 노원구 --dong 상계동 \
      --n 8 \
      --out .claude/skills/carousel/brands/zipsaja/reels/src/data/nowon-sanggye.json
"""
import argparse, json, os, sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl 필요: pip install openpyxl")


def load_rows(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return list(ws.iter_rows(min_row=2, values_only=True))


def curate(rows, gu: str, dong: str, n: int):
    filtered = [r for r in rows if r[0] == gu and r[1] == dong]
    if not filtered:
        sys.exit(f"{gu} {dong} 매물 없음")

    by_complex = defaultdict(list)
    for r in filtered:
        by_complex[r[2]].append(r)

    # 단지별 가격 범위 + 세대수
    complexes = []
    for name, items in by_complex.items():
        items_p = sorted(items, key=lambda x: x[8] if x[8] else 999)
        best = items_p[0]
        prices = [x[8] for x in items if x[8]]
        if not prices:
            continue
        complexes.append({
            "name": name,
            "households": best[3],
            "year": best[4],
            "pyeong": best[5],
            "priceMin": min(prices),
            "priceMax": max(prices),
            "desc": (best[14] or "").strip(),
            "count": len(items),
        })

    # 세대수 큰 순 상위 n개, 다시 가격순
    complexes.sort(key=lambda x: -(x["households"] or 0))
    picked = complexes[:n]
    picked.sort(key=lambda x: x["priceMin"])

    price_min = min(c["priceMin"] for c in picked)
    meta = {
        "gu": gu,
        "dong": dong,
        "priceTag": f"{int(price_min)}억대부터",
        "totalListings": len(filtered),
        "uniqueComplexes": len(by_complex),
    }
    return {"meta": meta, "complexes": picked}


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--excel", required=True)
    p.add_argument("--gu", required=True)
    p.add_argument("--dong", required=True)
    p.add_argument("--n", type=int, default=8)
    p.add_argument("--out", required=True)
    args = p.parse_args()

    rows = load_rows(args.excel)
    data = curate(rows, args.gu, args.dong, args.n)

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"✓ {args.out}")
    print(f"  {data['meta']['gu']} {data['meta']['dong']}: "
          f"{data['meta']['totalListings']}건, {len(data['complexes'])}개 단지 선별")


if __name__ == "__main__":
    main()
