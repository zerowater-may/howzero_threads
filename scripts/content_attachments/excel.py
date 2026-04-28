"""Excel generator using openpyxl — outputs a styled .xlsx with dataset rows."""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def write_excel(dataset: dict[str, Any], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = dataset.get("sheetTitle", "데이터")

    # A1: title (merged)
    ws["A1"] = dataset["title"]
    ws["A1"].font = Font(name="Malgun Gothic", size=16, bold=True, color="EA2E00")
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # A2: subtitle (period + source)
    ws["A2"] = f"{dataset['periodLabel']} · {dataset['source']}"
    ws["A2"].font = Font(name="Malgun Gothic", size=11, italic=True, color="555555")
    ws.merge_cells("A2:D2")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Row 3: column headers
    headers = [
        dataset.get("districtLabel", "지역"),
        f"{dataset.get('beforeLabel', '취임 전')} (만원)",
        f"{dataset.get('afterLabel', '취임 후')} (만원)",
        f"{dataset.get('changeLabel', '변동률')}(%)",
    ]
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=name)
        cell.font = Font(name="Malgun Gothic", size=12, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="EA2E00", end_color="EA2E00", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for i, d in enumerate(dataset["districts"], start=4):
        ws.cell(row=i, column=1, value=d["district"]).font = Font(name="Malgun Gothic", bold=True)
        ws.cell(row=i, column=2, value=d["priceBefore"])
        ws.cell(row=i, column=3, value=d["priceAfter"])
        pct_cell = ws.cell(row=i, column=4, value=d["changePct"])
        pct_cell.font = Font(
            name="Malgun Gothic",
            color="EA2E00" if d["changePct"] >= 0 else "1A4FA0",
            bold=True,
        )

    # Column widths
    for col, width in enumerate([12, 18, 18, 14], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
