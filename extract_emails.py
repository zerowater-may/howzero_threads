"""Threads 댓글에서 이메일 추출 → 중복 제거 → XLSX 저장."""

import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from howzero_threads.api.client import ThreadsClient
from howzero_threads.api.comments import get_comments
from howzero_threads.config import Settings

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")

MEDIA_ID = "17853491439596973"  # DVKgJ_nD8uh
OUTPUT = "data/comments_emails.xlsx"


def main():
    settings = Settings()
    client = ThreadsClient(settings)

    print(f"댓글 조회 중... (media_id: {MEDIA_ID})")
    comments = get_comments(client, MEDIA_ID, include_hidden=False)
    print(f"총 {len(comments)}건 댓글 수집 완료")

    # 이메일 기준 중복 제거 (첫 번째 댓글만 유지)
    seen_emails = set()
    rows = []
    for c in comments:
        text = c.get("text", "")
        emails = EMAIL_RE.findall(text)
        if emails:
            email = emails[0].lower()
            if email in seen_emails:
                continue
            seen_emails.add(email)
            rows.append({
                "username": c.get("username", ""),
                "email": email,
                "text": text,
                "timestamp": c.get("timestamp", ""),
                "permalink": c.get("permalink", ""),
                "comment_id": c.get("id", ""),
            })

    print(f"이메일 포함 댓글: {len(rows)}건 (중복 제거 후)")

    if not rows:
        print("이메일이 포함된 댓글이 없습니다.")
        sys.exit(0)

    # XLSX 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "댓글 이메일"

    headers = ["#", "Username", "Email", "Comment", "Timestamp", "Permalink"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(rows, 1):
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=r["username"])
        ws.cell(row=i + 1, column=3, value=r["email"])
        ws.cell(row=i + 1, column=4, value=r["text"])
        ws.cell(row=i + 1, column=5, value=r["timestamp"])
        ws.cell(row=i + 1, column=6, value=r["permalink"])

    # 열 너비 조정
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 50

    wb.save(OUTPUT)
    print(f"저장 완료: {OUTPUT}")

    # Gmail 복사용 이메일 리스트 출력
    print(f"\n{'='*60}")
    print("Gmail BCC용 이메일 (복사해서 붙여넣기):")
    print("="*60)
    all_emails = [r["email"] for r in rows]
    print(", ".join(all_emails))
    print(f"\n총 {len(all_emails)}명")


if __name__ == "__main__":
    main()
