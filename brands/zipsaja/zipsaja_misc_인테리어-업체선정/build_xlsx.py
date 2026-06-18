# -*- coding: utf-8 -*-
"""
인테리어 업체 선정 비교표 생성기 (파스텔톤 4시트)
- 시트1: 업체 한눈에 비교 (5곳)
- 시트2: 항목별 견적 비교 (정규화, 리바트/한샘 검산 일치)
- 시트3: 업체 선정 체크리스트 (5곳)
- 시트4: 상담 녹취 요약 (아파트멘터리 / 퍼스트에비뉴 / 홈루덴스)
- 시트5: 한샘 가견적 상세
- 시트6: 업체 미팅 질문 리스트
- 시트7: 홈루덴스 방문상담 상세 (2026-06-13 방문 녹음 정리)

실행: uv run --python 3.12 --with openpyxl build_xlsx.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "맑은 고딕"

# ---- 색 팔레트 (파스텔) ----
SLATE      = "7E8B99"   # 메인 타이틀 배경 (진한 슬레이트)
MINT       = "9FD8CB"   # 보조 타이틀
HEADER     = "DCE2E8"   # 표 헤더 회색
CREAM      = "F4F0E9"   # 카테고리/라벨 열
GRID       = "D9D9D9"   # 테두리
LOWEST     = "FFE49A"   # 최저가 강조 (파스텔 옐로)
WARN       = "FBE2E2"   # 주의 (연핑크)
GOOD       = "E2F0DD"   # 강점 (연그린)

LIVART_H,  LIVART_C  = "B6D7A8", "EAF4E6"   # 리바트  - 그린
HANS_H,    HANS_C    = "A9CCE3", "E8F1FB"   # 한샘    - 블루
APT_H,     APT_C     = "F5CBA7", "FCEEE3"   # 아파트멘터리 - 피치
FIRST_H,   FIRST_C   = "C9B6E4", "F1EAF9"   # 퍼스트에비뉴 - 라벤더
HOME_H,    HOME_C    = "A8DAD0", "E8F6F2"   # 홈루덴스 - 민트

def fill(hexc): return PatternFill("solid", fgColor=hexc)
def font(size=10, bold=False, color="2B2B2B"): return Font(name=FONT, size=size, bold=bold, color=color)

thin = Side(style="thin", color=GRID)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center", wrap_text=True)

def put(ws, r, c, value, *, fillc=None, f=None, align=WRAP, numfmt=None, border=True):
    cell = ws.cell(row=r, column=c, value=value)
    if fillc: cell.fill = fill(fillc)
    cell.font = f or font()
    cell.alignment = align
    if border: cell.border = BORDER
    if numfmt: cell.number_format = numfmt
    return cell

def title_row(ws, r, ncols, text, bg=SLATE, fg="FFFFFF", size=15, h=34):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1, value=text)
    c.fill = fill(bg); c.font = font(size, True, fg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = h

def note_row(ws, r, ncols, text, bg="F6F4EF", fg="7A6F5F", size=9, h=30):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1, value=text)
    c.fill = fill(bg); c.font = font(size, False, fg)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = h

WON = '#,##0"원"'

wb = Workbook()

# =====================================================================
# 시트 1 · 업체 한눈에 비교
# =====================================================================
ws = wb.active
ws.title = "1.업체비교"
ws.sheet_view.showGridLines = False
NC = 6
widths = [18, 25, 25, 25, 25, 25]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

title_row(ws, 1, NC, "인테리어 업체 비교  ·  봉천동 두산아파트 59A(24평형)  ·  전체+확장2+에어컨4+샷시 / 9월 공사")
note_row(ws, 2, NC, "견적서 2곳(리바트·한샘)은 정식 견적, 전화 3곳(아파트멘터리·퍼스트에비뉴·홈루덴스)은 상담 추정치입니다. 추정치는 방문/현장실측 시 변동됩니다.")

# 헤더
hr = 3
put(ws, hr, 1, "비교 항목", fillc=HEADER, f=font(11, True), align=CENTER)
put(ws, hr, 2, "리바트 (LIVART)",  fillc=LIVART_H, f=font(11, True), align=CENTER)
put(ws, hr, 3, "한샘 (HANSSEM)",   fillc=HANS_H,   f=font(11, True), align=CENTER)
put(ws, hr, 4, "아파트멘터리",      fillc=APT_H,    f=font(11, True), align=CENTER)
put(ws, hr, 5, "퍼스트에비뉴",      fillc=FIRST_H,  f=font(11, True), align=CENTER)
put(ws, hr, 6, "홈루덴스",          fillc=HOME_H,   f=font(11, True), align=CENTER)
ws.row_dimensions[hr].height = 26

# (라벨, 리바트, 한샘, 아파트멘터리, 퍼스트, 홈루덴스, 행높이, 특수)
rows1 = [
    ("견적 형태",        "정식 견적서 (종이)", "정식 견적서 (엑셀)", "전화 상담", "전화 상담", "전화+방문상담 완료\n(6/13, 2시간)", 30, None),
    ("견적 총액(VAT포함)", 55255690, 78127585, "7,000~7,500만원+\n(가구 별도)", "약 7,800만원\n(상담 추정)", "고객 상한 6천 초반\n(VAT포함·2차서 견적)", 40, "money"),
    ("평당가(24평 환산)",  2302000, 3255000, "약 290만+ (추정)", "약 325만 (추정)", "약 250만 (6천÷24)", 24, "money2"),
    ("담당 / 연락처",     "강정두\n010-5227-0351", "(견적서 미표기)", "압구정 본사\n(방문상담 권유)", "(상담원)", "방문 디자이너\n2차미팅 6/18(목) 19시", 38, None),
    ("소재지",           "-", "-", "압구정 본사", "강남 논현동", "서초(남부터미널)·목동", 24, None),
    ("공사 기간",        "3주 예상", "(미표기)", "잔금 8/28 후 9월 시작", "9월", "8월말 착공~9/30 입주", 26, None),
    ("A/S 보증",         "무상 3년", "견적서 명시 (무상)", "(미확인)", "1년 무상", "(2차 미팅서 확인)", 24, None),
    ("창호(샷시)",       "현대 L&C 이중창\n1,300만원", "1,118만원", "전체 교체\n(브랜드 미언급)", "KCC", "HOMECC(홈씨씨) 확정\n가스켓 방식·곰팡이↓", 40, None),
    ("주방 / 가구 방식",  "리바트 M100 (자체)", "한샘 주방·수납", "전체 맞춤제작\n(한샘·리바트 안 씀)", "주방 한솔 / 사양 맞춤", "한샘(부엌·현관장)\n욕실장은 제작", 40, None),
    ("붙박이장",         "리바트 M100\n(견적 공란)", "수납공사 0원\n(확인 필요)", "별도\n(한샘·리바트 별도)", "(미언급)", "안방 수납 일단 제외\n짜기/빼기 견적 둘 다", 40, None),
    ("발코니 확장",       "포함 (거실+방, 2개소)", "포함", "포함 (2개소)", "포함 (2개소)", "포함 (거실+부엌옆방)\n입구방 제외(세탁기)", 40, None),
    ("에어컨",           "시스템 단내림 2개소\n(목공 포함)", "환기/공조 800만", "시스템 4대 포함", "삼성 무풍 4대", "견적 제외→직접결제\n(4대, 단내림 목공 포함)", 40, None),
    ("강점",            "최저가 · 직영시공 · 3년 A/S", "대기업 풀패키지 · 자재 구성 명확", "맞춤제작 디자인 · 사례 풍부", "자재 브랜드 명확 · 사례 기반", "방문 디테일 압도적 · 6천내 맞춤 · 에어컨/공산품 분리 투명 · HOMECC 가스켓", 50, "good"),
    ("주의 / 약점",      "TV장·붙박이장 견적 공란\n입주청소·승강기 별도", "최고가\n수납 0원 항목 확인", "가구 별도라 총액 상승\n방문상담 필요", "전화 추정 견적\nA/S 1년으로 짧음", "2차 견적서 아직\n에어컨·가구 별도\nA/S 미확인", 48, "warn"),
    ("자료 출처",        "리바트_견적서.png", "한샘_견적서.png", "아파트멘터리_전사.txt", "퍼스트에비뉴_전사.txt", "홈루덴스_방문상담록.md\n(+전사·전화)", 24, None),
]

r = hr + 1
TOTAL_ROW = None
for label, v_li, v_ha, v_ap, v_fi, v_ho, h, kind in rows1:
    put(ws, r, 1, label, fillc=CREAM, f=font(10, True), align=Alignment(horizontal="left", vertical="center", wrap_text=True))
    cells = [(2, v_li, LIVART_C), (3, v_ha, HANS_C), (4, v_ap, APT_C), (5, v_fi, FIRST_C), (6, v_ho, HOME_C)]
    for col, val, cbg in cells:
        if kind == "money" and isinstance(val, int):
            put(ws, r, col, val, fillc=cbg, align=RIGHT, numfmt=WON, f=font(11, True))
        elif kind == "money2" and isinstance(val, int):
            put(ws, r, col, val, fillc=cbg, align=RIGHT, numfmt=WON)
        elif kind == "good":
            put(ws, r, col, val, fillc=GOOD)
        elif kind == "warn":
            put(ws, r, col, val, fillc=WARN)
        else:
            al = RIGHT if (isinstance(val, str) and ("만원" in val or "만+" in val)) else WRAP
            put(ws, r, col, val, fillc=cbg, align=al)
    if kind == "money":
        TOTAL_ROW = r
    ws.row_dimensions[r].height = h
    r += 1

# 최저가(리바트) 총액 강조
if TOTAL_ROW:
    ws.cell(row=TOTAL_ROW, column=2).fill = fill(LOWEST)
    ws.cell(row=TOTAL_ROW, column=2).font = font(11, True, "9E7B4E")

ws.freeze_panes = "B4"

# =====================================================================
# 시트 2 · 항목별 견적 비교 (정규화) — 견적서 받은 리바트/한샘만
# =====================================================================
ws2 = wb.create_sheet("2.항목별견적")
ws2.sheet_view.showGridLines = False
NC2 = 7
w2 = [16, 13, 13, 14, 14, 13, 40]
for i, w in enumerate(w2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

title_row(ws2, 1, NC2, "항목별 견적 비교 (5개 업체)", bg=SLATE)
note_row(ws2, 2, NC2,
         "리바트·한샘은 받은 견적서를 공통 카테고리로 정규화해 채웠습니다(합계 검산 일치: 리바트 55,255,690 / 한샘 78,127,585). "
         "전화 상담 3곳(아파트멘터리·퍼스트에비뉴·홈루덴스)은 견적서를 받으면 빈칸에 직접 입력하세요 — 숫자만 적으면 '원' 단위가 자동 표시됩니다. 맨 오른쪽 비고는 리바트·한샘의 원본 항목명입니다.",
         h=42)

hr2 = 3
heads2 = ["공통 카테고리", "리바트", "한샘", "아파트멘터리", "퍼스트에비뉴", "홈루덴스", "원본 항목 (리바트 · 한샘)"]
hcol = [HEADER, LIVART_H, HANS_H, APT_H, FIRST_H, HOME_H, HEADER]
for i, (t, cbg) in enumerate(zip(heads2, hcol), 1):
    put(ws2, hr2, i, t, fillc=cbg, f=font(10, True), align=CENTER)
ws2.row_dimensions[hr2].height = 30

# (카테고리, 리바트액, 리바트비고, 한샘액, 한샘비고)
rows2 = [
    ("철거",            2350000, "몰딩·가구·바닥 철거 + 폐기물처리",            3693076, "철거공사(시공비)"),
    ("설비",            0,       "확장공사에 포함",                            5122775, "설비공사(시공비)"),
    ("목공",            4200000, "문틀교체(방3/욕1)·에어컨단내림·마이너스몰딩", 8096501, "목공사(시공비)"),
    ("확장(발코니)",     5900000, "거실+방 발코니 확장 (설비+목공)",            0,       "별도 항목 없음 (목공/설비에 포함 추정)"),
    ("전기",            3550000, "인덕션선·차단기·기본등·매립등·스위치·인건비", 3451606, "전기공사(시공비)"),
    ("조명",            0,       "전기 '기본등'에 포함",                       2191375, "조명공사 441,375 + 조명자재 1,750,000"),
    ("타일",            1700000, "주방·베란다·현관 (덧방)",                    3883415, "타일공사(시공비)"),
    ("도배",            3000000, "실크 전체 + 슬림문틀 퍼티",                  3002741, "도배공사(시공비)"),
    ("도장(칠)",        900000,  "발코니·세탁실 세라믹탄성",                   779816,  "도장공사(시공비)"),
    ("필름",            250000,  "현관문 내부면",                             213465,  "필름공사(시공비)"),
    ("주방가구",         5000000, "씽크대 리바트 M100 (아일랜드 포함, 쿡탑/식세기 별도)", 6500000, "주방공사(시공+자재)"),
    ("수납·붙박이장",    850000,  "현관장. TV장·안방붙박이장은 견적 공란",       0,       "수납공사 0원 (확인 필요)"),
    ("욕실",            4900000, "공용욕실 전체철거 (포세린 +70만)",            3822000, "바스공사(시공+자재)"),
    ("창호(샷시)",       13000000,"현대 L&C 이중창/안방",                       11181246,"창호공사(시공+자재)"),
    ("마루·바닥",        3600000, "강마루 4각 (포세린타일 +140만)",             6485000, "마루/장판(시공+자재)"),
    ("중문·도어·몰딩",   0,       "목공 터닝도어에 포함",                       2919200, "중문/도어/몰딩(자재)"),
    ("환기·공조",        0,       "해당 없음 (에어컨 단내림은 목공)",            8000000, "환기/공조 설치(시공+자재)"),
    ("마감자재",         0,       "위 항목에 포함",                            6087900, "벽지·필름·벽장재·타일 자재"),
    ("준공·입주청소",    0,       "입주청소 별도",                             1723770, "공사마감 및 준공청소"),
    ("현장관리·감리",    2927900, "현장진행비 1,748,000 + 직영감리 1,179,900",  0,       "별도 명시 없음"),
    ("기타",            650000,  "가스관·보양·동의서 (VAT별도)",               973700,  "기타공사_한샘시공 외"),
    ("부가세",          2477790, "1차 기본공사분 10% (가구는 VAT포함가)",      0,       "합계에 VAT 포함"),
]

BLANK = "FFFFFF"   # 견적서 미수령 → 직접 입력할 빈칸
r = hr2 + 1
for cat, li, li_n, ha, ha_n in rows2:
    put(ws2, r, 1, cat, fillc=CREAM, f=font(10, True))
    put(ws2, r, 2, li if li else "-", fillc=LIVART_C, align=RIGHT, numfmt=WON if li else None)
    put(ws2, r, 3, ha if ha else "-", fillc=HANS_C, align=RIGHT, numfmt=WON if ha else None)
    # 견적서 안 받은 3곳 → 빈칸(흰색). 숫자 입력 시 '원' 자동 표시되도록 numfmt 미리 지정
    put(ws2, r, 4, None, fillc=BLANK, align=RIGHT, numfmt=WON)
    put(ws2, r, 5, None, fillc=BLANK, align=RIGHT, numfmt=WON)
    put(ws2, r, 6, None, fillc=BLANK, align=RIGHT, numfmt=WON)
    note = f"리바트: {li_n}\n한샘: {ha_n}"
    put(ws2, r, 7, note, fillc="F7F7F4", f=font(9, color="555555"))
    ws2.row_dimensions[r].height = 34
    r += 1

# 합계행
put(ws2, r, 1, "총 계 (VAT포함)", fillc=SLATE, f=font(11, True, "FFFFFF"))
put(ws2, r, 2, 55255690, fillc=LOWEST, align=RIGHT, numfmt=WON, f=font(11, True, "9E7B4E"))
put(ws2, r, 3, 78127585, fillc=HANS_H, align=RIGHT, numfmt=WON, f=font(11, True))
put(ws2, r, 4, None, fillc=BLANK, align=RIGHT, numfmt=WON)
put(ws2, r, 5, None, fillc=BLANK, align=RIGHT, numfmt=WON)
put(ws2, r, 6, None, fillc=BLANK, align=RIGHT, numfmt=WON)
put(ws2, r, 7, "전화 상담 3곳은 견적서 수령 후 위 빈칸과 이 합계칸에 직접 입력", fillc="F7F7F4", f=font(9, color="888888"))
ws2.row_dimensions[r].height = 28
ws2.freeze_panes = "B4"

# =====================================================================
# 시트 3 · 업체 선정 체크리스트
# =====================================================================
ws3 = wb.create_sheet("3.선정체크리스트")
ws3.sheet_view.showGridLines = False
NC3 = 9
w3 = [13, 30, 7, 15, 15, 15, 15, 15, 20]
for i, w in enumerate(w3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

title_row(ws3, 1, NC3, "업체 선정 체크리스트 (견적서·통화에서 발견한 실제 확인 포인트)", bg=SLATE)
note_row(ws3, 2, NC3, "중요도: 필수 / 권장 / 참고. 빈 '내 확인/메모' 칸은 업체에 직접 물어보고 채우세요. 견적서·통화에서 확인된 내용은 미리 적어 뒀습니다.")

hr3 = 3
heads3 = ["카테고리", "확인 항목", "중요도", "리바트", "한샘", "아파트멘터리", "퍼스트에비뉴", "홈루덴스", "내 확인 / 메모"]
hcol3 = [HEADER, HEADER, HEADER, LIVART_H, HANS_H, APT_H, FIRST_H, HOME_H, CREAM]
for i, (t, cbg) in enumerate(zip(heads3, hcol3), 1):
    put(ws3, hr3, i, t, fillc=cbg, f=font(10, True), align=CENTER)
ws3.row_dimensions[hr3].height = 28

# (카테고리, 항목, 중요도, 리바트, 한샘, 아파트멘터리, 퍼스트, 홈루덴스, 메모)
rows3 = [
    ("견적 범위", "폐기물 처리 포함?",              "권장",  "포함(60만)", "철거에 포함 추정", "확인 필요", "확인 필요", "확인 필요", ""),
    ("견적 범위", "입주청소 포함?",                "권장",  "별도", "준공청소 포함", "확인 필요", "확인 필요", "확인 필요", ""),
    ("견적 범위", "승강기 사용료 / 보양",          "참고",   "보양 35만, 사용료 별도", "확인 필요", "확인 필요", "확인 필요", "확인 필요", ""),
    ("견적 범위", "주민 공사동의서 처리",          "참고",   "15만 포함", "확인 필요", "확인 필요", "확인 필요", "확인 필요", ""),
    ("견적 범위", "철거 범위 (가구·바닥·몰딩)",     "권장",  "명시됨", "시공비 일괄", "확인 필요", "확인 필요", "확인 필요", ""),
    ("가격 조건", "부가세 포함 여부",              "필수", "포함", "포함", "확인 필요", "확인 필요", "별도 조정 가능", ""),
    ("가격 조건", "추가금 발생 조건 (실측 후 변동)", "필수", "REMARK 명시", "특이사항 명시", "방문상담서 확정", "현장실측 후", "방문상담서 확정", ""),
    ("가격 조건", "가구·붙박이장 별도 여부",        "필수", "붙박이장 견적 공란", "수납 0원", "별도(가구 미포함)", "주방만 포함 추정", "에어컨 별도 권장", ""),
    ("가격 조건", "할인 / 단수정리 여지",          "참고",   "확인 필요", "단수정리 있음", "확인 필요", "확인 필요", "예산 5천 맞춤", ""),
    ("자재 사양", "창호 브랜드 / 등급",            "필수", "현대 L&C 이중창", "브랜드 확인 필요", "확인 필요", "KCC", "LX/KCC/HOMECC", ""),
    ("자재 사양", "마루 종류",                    "권장",  "강마루 4각", "마루/장판", "확인 필요", "구정마루 강마루", "확인 필요", ""),
    ("자재 사양", "타일 (덧방/철거, 등급)",        "권장",  "덧방+현관 철거후", "타일공사", "확인 필요", "확인 필요", "확인 필요", ""),
    ("자재 사양", "도배 (실크/합지)",              "권장",  "실크 전체", "확인 필요", "확인 필요", "디아망(공용부)", "확인 필요", ""),
    ("자재 사양", "주방 가구 브랜드",              "권장",  "리바트 M100", "한샘", "맞춤제작", "한솔/맞춤", "한샘 위주", ""),
    ("자재 사양", "에어컨 (대수/방식)",            "권장",  "단내림 2개소", "환기/공조", "4대", "삼성무풍 4대", "별도(직접결제 권장)", ""),
    ("시공",     "직영 시공 vs 하도급",           "필수", "직영(M&S)", "확인 필요", "확인 필요", "확인 필요", "확인 필요", ""),
    ("시공",     "공사 기간",                    "권장",  "3주", "확인 필요", "확인 필요", "확인 필요", "8말~9/30", ""),
    ("시공",     "감리 / 현장소장 유무",          "권장",  "감리 5% 명시", "확인 필요", "확인 필요", "확인 필요", "확인 필요", ""),
    ("시공",     "9월 일정 가능 여부 (잔금 8/28)", "필수", "확인 필요", "확인 필요", "8/28후 가능", "9월 가능", "9/30 입주 맞춤", ""),
    ("보증",     "A/S 기간",                     "필수", "무상 3년", "무상(명시)", "확인 필요", "1년 무상", "확인 필요", ""),
    ("보증",     "하자보수 무상 범위",            "권장",  "소모품 별도", "확인 필요", "확인 필요", "확인 필요", "확인 필요", ""),
    ("계약",     "표준 계약서 작성 여부",          "필수", "확인 필요", "산식소 계약서", "확인 필요", "확인 필요", "확인 필요", ""),
    ("계약",     "대금 지급 일정 (계약/중도/잔금)", "필수", "확인 필요", "확인 필요", "확인 필요", "확인 필요", "확인 필요", ""),
    ("계약",     "추가공사 서면 문서화",           "권장",  "확인 필요", "추후 계약서로", "확인 필요", "확인 필요", "확인 필요", ""),
]

def imp_color(s):
    return "BC8079" if s == "필수" else ("C39A4A" if s == "권장" else "888888")

cat_fill_cycle = {"견적 범위": "FBF7EF", "가격 조건": "F0F5FA", "자재 사양": "F1F8EE", "시공": "FAF1F7", "보증": "F3F0FA", "계약": "FDF3E9"}
r = hr3 + 1
for cat, item, imp, li, ha, ap, fi, ho, memo in rows3:
    cbg = cat_fill_cycle.get(cat, "FFFFFF")
    put(ws3, r, 1, cat, fillc=CREAM, f=font(9, True))
    put(ws3, r, 2, item, fillc=cbg, f=font(10))
    put(ws3, r, 3, imp, fillc=cbg, align=CENTER, f=font(10, True, imp_color(imp)))
    put(ws3, r, 4, li, fillc=LIVART_C, f=font(9), align=CENTER)
    put(ws3, r, 5, ha, fillc=HANS_C, f=font(9), align=CENTER)
    put(ws3, r, 6, ap, fillc=APT_C, f=font(9), align=CENTER)
    put(ws3, r, 7, fi, fillc=FIRST_C, f=font(9), align=CENTER)
    put(ws3, r, 8, ho, fillc=HOME_C, f=font(9), align=CENTER)
    put(ws3, r, 9, memo, fillc="FFFFFF")
    ws3.row_dimensions[r].height = 26
    r += 1
ws3.freeze_panes = "B4"

# =====================================================================
# 시트 4 · 상담 녹취 요약
# =====================================================================
ws4 = wb.create_sheet("4.상담녹취요약")
ws4.sheet_view.showGridLines = False
NC4 = 4
w4 = [18, 33, 33, 33]
for i, w in enumerate(w4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

title_row(ws4, 1, NC4, "전화 상담 녹취 요약", bg=SLATE)
note_row(ws4, 2, NC4, "통화 녹음을 전사해 핵심만 정리했습니다. 전체 원문은 통화_전사/ 폴더의 txt를 참고하세요. 추정 견적은 방문상담·실측 시 변동됩니다.")

hr4 = 3
put(ws4, hr4, 1, "항목", fillc=HEADER, f=font(11, True), align=CENTER)
put(ws4, hr4, 2, "아파트멘터리", fillc=APT_H, f=font(11, True), align=CENTER)
put(ws4, hr4, 3, "퍼스트에비뉴", fillc=FIRST_H, f=font(11, True), align=CENTER)
put(ws4, hr4, 4, "홈루덴스", fillc=HOME_H, f=font(11, True), align=CENTER)
ws4.row_dimensions[hr4].height = 26

rows4 = [
    ("통화 일시",      "2026-06-11 14:47", "2026-06-11 16:18", "전화 06-11 18:33\n방문 06-13 (2시간)", 32),
    ("우리집 정보",    "봉천동 두산아파트 107동 1202호\n24평형 24A타입", "봉천동 주산아파트\n59A타입(=24평형)", "A타입 (좌측 거실)\n거주는 삼성동", 38),
    ("공사 범위",      "전체공사 · 발코니 2개소(거실+작은방) 확장\n시스템에어컨 4대 · 욕실 1개소 · 주방", "전체공사 · 베란다 2개소 확장\n에어컨 4대 · 샷시 전체교체", "전체 창호교체 + 확장 2개소\n(거실+부엌옆방, 입구방 제외)\n에어컨 4대 · 욕실 1", 50),
    ("추정 견적",      "미니멈 7,000~7,500만원~\n(붙박이장 양에 따라 상승)", "약 7,800만원\n내부5,700+샷시900+확장500+에어컨700", "예산 5,000만원대 목표\n(에어컨 별도, VAT별도 조정)", 46),
    ("가구 방식",      "전체 맞춤제작 (한샘·리바트 안 씀)\n주방·신발장 자체 / 붙박이장 별도 가능", "주방 한솔 또는 고객 사양 맞춤", "한샘 위주 (선호 브랜드 가능)", 42),
    ("브랜드 / 자재",  "(방문상담서 세부 안내)", "샷시 KCC · 에어컨 삼성무풍\n바닥 구정마루 강마루 · 도배 디아망", "창호 LX/KCC/HOMECC\n(KCC계열 완성창 제안)", 42),
    ("A/S",           "(미확인)", "1년 무상", "(미확인)", 22),
    ("위치 / 다음 스텝","압구정 본사 방문상담 → 세부견적", "재상담 예정", "서초(남부터미널) 사무실\n토요일 방문상담 예약", 32),
    ("핵심 포인트",    "참고 7천만원에 가구(붙박이장) 미포함\n참고 맞춤제작이라 디자인 자유도 높음\n참고 방문상담 필수", "참고 자재 브랜드가 명확하고 사례 기반\n참고 A/S 1년으로 견적서 2곳(3년)보다 짧음\n참고 전화 추정이라 실측 필요", "참고 6/13 방문상담 완료→'7.홈루덴스_방문상담' 시트\n참고 에어컨 직접결제(마진절감)\n참고 창호 HOMECC 확정·가구 한샘\n참고 2차미팅 6/18(목) 19시 (서초)", 62),
]

r = hr4 + 1
for label, ap, fi, ho, h in rows4:
    put(ws4, r, 1, label, fillc=CREAM, f=font(10, True))
    hl = "핵심" in label
    put(ws4, r, 2, ap, fillc=GOOD if hl else APT_C, f=font(10, hl))
    put(ws4, r, 3, fi, fillc=GOOD if hl else FIRST_C, f=font(10, hl))
    put(ws4, r, 4, ho, fillc=GOOD if hl else HOME_C, f=font(10, hl))
    ws4.row_dimensions[r].height = h
    r += 1
ws4.freeze_panes = "B4"

# =====================================================================
# 시트 5 · 한샘 가견적 상세 (하위 카테고리)
# =====================================================================
ws5 = wb.create_sheet("5.한샘가견적상세")
ws5.sheet_view.showGridLines = False
NC5 = 11
w5 = [15, 32, 6, 6, 12, 13, 12, 12, 12, 12, 22]
for i, w in enumerate(w5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

title_row(ws5, 1, NC5, "한샘 가견적 상세 (하위 카테고리, 추후 변동)", bg=SLATE)
note_row(ws5, 2, NC5,
         "한샘 가견적 6장을 카테고리별로 펼쳤습니다. 단가·수량·시공비·소계는 검산으로 확인했습니다. "
         "사진상 품명(세부 항목명) 왼쪽이 잘린 페이지가 많아, 벽지·타일자재와 환기공조를 제외하면 품명은 비어 있습니다(엑셀 원본을 받으면 품명까지 채울 수 있음). "
         "리바트는 카테고리 소계만 비교용으로 넣었고, 전화 상담 3곳은 빈칸입니다.", h=54)

hr5 = 3
heads5 = ["공정", "세부 품명", "단위", "수량", "단가", "한샘 시공비", "리바트", "아파트멘터리", "퍼스트에비뉴", "홈루덴스", "비고"]
hcol5 = [HEADER, HEADER, HEADER, HEADER, HEADER, HANS_H, LIVART_H, APT_H, FIRST_H, HOME_H, HEADER]
for i, (t, cbg) in enumerate(zip(heads5, hcol5), 1):
    put(ws5, hr5, i, t, fillc=cbg, f=font(10, True), align=CENTER)
ws5.row_dimensions[hr5].height = 28

# 카테고리: (공정명, 한샘 소계, 리바트 소계, [세부행(품명, 단위, 수량, 단가, 시공비)], 비고)
cats = [
    ("1. 철거공사", 3693076, 2350000, [
        ("", "M", 12, 17900, 214800), ("", "EA", 1, 27500, 27500), ("", "지", 5, 17900, 89500),
        ("", "평", 12, 34400, 412800), ("", "EA", 3, 57500, 172500), ("", "평", 25, 5160, 129000),
        ("", "평", 25, 5160, 129000), ("몰딩·문짝·문선·인방 포함", "SET", 4, 44300, 177200),
        ("", "M", 10, 19300, 193000), ("", "㎡", 5, 22000, 110000), ("", "㎡", 0, 19300, 0),
        ("바닥 낮춤(철거)", "M", 1, 68800, 68800), ("", "㎡", 0, 19300, 0), ("", "EA", 10, 22000, 220000),
        ("", "EA", 1, 921900, 921900), ("(철거)", "EA", 1, 147500, 147500),
        ("폐기물(철거자재·비닐·가위류·블라인드 등)", "EA", 1, 452025, 452025),
        ("폐기물 처리비(5%)", "", 1, 227551, 227551),
    ], "검산 일치 · 품명 일부 잘림"),
    ("2. 설비공사", 4774383, 0, [
        ("", "EA", 1, 199300, 199300), ("", "EA", 1, 325000, 325000), ("", "EA", 1, 161600, 161600),
        ("", "EA", 1, 1894700, 1894700), ("단열공사", "EA", 1, 1475800, 1475800),
        ("설비 시공비(10%)", "", 1, 405640, 405640), ("", "", 1, 312343, 312343),
    ], "검산 일치 · 품명 잘림"),
    ("3. 목공공사", 8096501, 4200000, [
        ("", "평", 25, 22000, 550000), ("", "EA", 1, 824600, 824600), ("", "EA", 3, 693500, 2080500),
        ("", "EA", 25, 11700, 292500), ("", "EA", 4, 61750, 247000), ("", "EA", 4, 31600, 126400),
        ("", "EA", 4, 44240, 176960), ("기준몰딩 등", "EA", 4, 146600, 586400), ("", "EA", 1, 305100, 305100),
        ("1py+자재", "㎡", 1, 354000, 354000), ("공사진행 30%", "", 1, 1663038, 1663038),
        ("시공비 5%", "", 1, 360325, 360325), ("공사진행비 7%", "", 1, 529678, 529678),
    ], "검산 일치 · 품명 잘림"),
    ("4. 전기공사", 3451606, 3550000, [
        ("", "EA", 5, 44500, 223500), ("", "EA", 5, 26500, 132500), ("콘센트(추가)", "EA", 20, 25100, 502000),
        ("4세대 등", "EA", 1, 282500, 282500), ("실설치", "EA", 1, 385400, 385400), ("", "EA", 1, 383600, 383600),
        ("", "EA", 6, 33700, 202200), ("", "EA", 1, 62100, 62100), ("", "EA", 4, 136000, 544000),
        ("시공비 20%", "", 1, 543580, 543580), ("공사진행비 7%", "", 1, 190246, 190246),
    ], "검산 일치 · 품명 잘림"),
    ("5. 조명공사", 441375, 0, [], "소계만 (세부 품명 미확인)"),
    ("6. 타일공사", 3883415, 1700000, [
        ("", "㎡", 10, 38500, 385000), ("", "㎡", 10, 13800, 138000), ("", "㎡", 5, 42600, 213000),
        ("", "㎡", 5, 13800, 69000), ("", "평", 10, 42600, 426000), ("", "㎡", 10, 13800, 138000),
        ("", "평", 25, 38500, 962500), ("", "㎡", 10, 42600, 426000), ("", "㎡", 35, 13800, 483000),
        ("", "", 1, 388860, 388860), ("부자재 포함", "EA", 1, 254055, 254055),
    ], "검산 일치 · 품명 잘림"),
    ("7. 도배공사", 3002241, 3000000, [
        ("", "EA", 25, 87500, 2187500), ("부자재 포함", "EA", 1, 616800, 616800),
        ("공사진행비 7%", "", 1, 196441, 196441),
    ], "검산 근사 · 품명 잘림"),
    ("8. 도장공사", 779816, 900000, [
        ("", "EA", 1, 550000, 550000), ("", "", 1, 178000, 178000), ("", "", 1, 53016, 53016),
    ], "검산 근사 · 품명 잘림"),
    ("9. 필름공사", 213465, 250000, [], "소계만 (세부 품명 미확인)"),
    ("10. 마감·준공청소", 1723770, 0, [], "요약견적 기준 (세부 미확인)"),
    ("11. 기타공사", 973700, 650000, [], "요약견적 기준 (세부 미확인)"),
    ("[스타일링] 주방공사", 6500000, 5000000, [], "소계만 (세부 품명 미확인)"),
    ("[스타일링] 수납공사", 0, 850000, [], "가견적 0원 (확인 필요)"),
    ("[스타일링] 바스(욕실)", 3822000, 4900000, [
        ("기본", "", 1, 2900000, 2900000), ("변기", "", 1, 150000, 150000),
        ("샤워기", "", 1, 350000, 350000), ("니플 고급형 SET", "", 1, 422000, 422000),
    ], "검산 일치 · 품명 일부 확인"),
    ("[스타일링] 중문·도어·몰딩", 2919200, 0, [], "소계만 (세부 품명 미확인)"),
    ("[스타일링] 창호공사", 11181246, 13000000, [
        ("창호 메인", "", 1, 9500000, 9500000), ("", "", 1, 831246, 831246),
        ("", "EA", 1, 400000, 400000), ("", "EA", 1, 450000, 450000),
    ], "검산 일치 · 품명 일부 잘림"),
    ("[스타일링] 마루·장판", 6485000, 3600000, [], "소계만 (세부 품명 미확인)"),
    ("[스타일링] 조명(자재)", 1750000, 0, [
        ("조명(자재)", "", 1, 1750000, 1750000),
    ], "7340 확인"),
    ("[스타일링] 벽지·필름·벽장재·타일(자재)", 6087900, 0, [
        ("실크벽지 (신한벽지 파스토/휠가드)", "EA", 25, 70900, 1772500),
        ("실크 전장지", "EA", 9, 42900, 386100),
        ("자정 필름 10M 1통", "", 1, 156500, 156500),
        ("현관타일 600*600 F_HSE타일", "", 4, 45200, 180800),
        ("주방타일 600*600 F_HSE타일", "", 10, 45200, 452000),
        ("욕실타일 600*600 F_HSE타일", "", 30, 45200, 1356000),
        ("발코니타일 300*300 F_HSE타일", "", 12, 32000, 384000),
        ("타일 부자재 (줄눈·압착·세라픽스류, 양생비 포함)", "", 56, 25000, 1400000),
    ], "검산 일치 · 품명 확인됨"),
    ("[스타일링] 환기·공조", 8000000, 0, [
        ("천정형 시스템 AC (삼성 시스템 AC)", "", 1, 8000000, 8000000),
    ], "품명 확인됨"),
]

QTY = '#,##0'
r = hr5 + 1
hanssem_total = 0
for name, sub_ha, sub_li, details, memo in cats:
    hanssem_total += sub_ha
    # 카테고리 소계행
    put(ws5, r, 1, name, fillc="E3E8EE", f=font(10, True, "33414F"))
    put(ws5, r, 2, "― 소계 ―", fillc="E3E8EE", f=font(9, True, "7A8794"), align=CENTER)
    put(ws5, r, 3, "", fillc="E3E8EE")
    put(ws5, r, 4, "", fillc="E3E8EE")
    put(ws5, r, 5, "", fillc="E3E8EE")
    put(ws5, r, 6, sub_ha, fillc=HANS_H, align=RIGHT, numfmt=WON, f=font(10, True))
    put(ws5, r, 7, sub_li if sub_li else None, fillc=LIVART_C, align=RIGHT, numfmt=WON if sub_li else None)
    put(ws5, r, 8, None, fillc=BLANK, align=RIGHT, numfmt=WON)   # 아파트멘터리 빈칸
    put(ws5, r, 9, None, fillc=BLANK, align=RIGHT, numfmt=WON)   # 퍼스트 빈칸
    put(ws5, r, 10, None, fillc=BLANK, align=RIGHT, numfmt=WON)  # 홈루덴스 빈칸
    put(ws5, r, 11, memo, fillc="E3E8EE", f=font(9, color="7A8794"))
    ws5.row_dimensions[r].height = 22
    r += 1
    # 세부행
    for pname, unit, qty, price, cost in details:
        put(ws5, r, 1, "", fillc="FFFFFF")
        put(ws5, r, 2, pname if pname else "(품명 잘림)", fillc="FFFFFF",
            f=font(9) if pname else font(9, color="BBBBBB"))
        put(ws5, r, 3, unit, fillc="FFFFFF", align=CENTER, f=font(9))
        put(ws5, r, 4, qty if isinstance(qty, int) else None, fillc="FFFFFF", align=CENTER, numfmt=QTY, f=font(9))
        put(ws5, r, 5, price if price else None, fillc="FFFFFF", align=RIGHT, numfmt=WON if price else None, f=font(9))
        put(ws5, r, 6, cost if cost else 0, fillc="FBFCFD", align=RIGHT, numfmt=WON, f=font(9))
        put(ws5, r, 7, None, fillc=BLANK)
        put(ws5, r, 8, None, fillc=BLANK)
        put(ws5, r, 9, None, fillc=BLANK)
        put(ws5, r, 10, None, fillc=BLANK)
        put(ws5, r, 11, "", fillc="FFFFFF")
        ws5.row_dimensions[r].height = 19
        r += 1

# 총합행
put(ws5, r, 1, "세부 합산 총계", fillc=SLATE, f=font(11, True, "FFFFFF"))
ws5.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
c = ws5.cell(row=r, column=2, value="(읽은 카테고리 합산 · 가견적)")
c.fill = fill(SLATE); c.font = font(9, False, "D6DDE4"); c.alignment = CENTER; c.border = BORDER
put(ws5, r, 6, hanssem_total, fillc=LOWEST, align=RIGHT, numfmt=WON, f=font(11, True, "9E7B4E"))
for cc in (7, 8, 9, 10):
    put(ws5, r, cc, None, fillc=BLANK, numfmt=WON)
put(ws5, r, 11, "요약 견적 78,127,585과 일부 차이(가견적)", fillc=SLATE, f=font(9, color="D6DDE4"))
ws5.row_dimensions[r].height = 26
ws5.freeze_panes = "B4"

# =====================================================================
# 시트 6 · 업체 미팅 질문 체크리스트 (현장에서 그대로 물어보기)
# =====================================================================
ws6 = wb.create_sheet("6.업체미팅_질문리스트")
ws6.sheet_view.showGridLines = False
NC6 = 6
w6 = [13, 42, 40, 8, 26, 4]
for i, w in enumerate(w6[:5], 1):
    ws6.column_dimensions[get_column_letter(i)].width = w

title_row(ws6, 1, 5, "업체 미팅 질문 체크리스트  (출처: 박목수의 열린견적서 안내서)", bg=SLATE)
note_row(ws6, 2, 5, "미팅에서 '이렇게 물어보세요' 칸을 그대로 물어보면 됩니다. '우리집 필수'로 표시한 항목은 우리집(봉천동 두산 59A·확장 2개소·샷시 전체교체·욕실 1개·에어컨 4대·9월 입주) 관련 질문입니다. 답변·메모 칸은 현장에서 적으세요.")

hr6 = 3
heads6 = ["분류", "이렇게 물어보세요", "왜 중요한지 / 주의", "중요도", "답변·메모"]
for i, t in enumerate(heads6, 1):
    put(ws6, hr6, i, t, fillc=HEADER, f=font(10, True), align=CENTER)
ws6.row_dimensions[hr6].height = 26

# 단계별 (제목, 헤더색, 띠색, [ (분류, 질문멘트, 이유, 중요도) ])
sections6 = [
    ("1단계. 견적 받을 때", "F5CBA7", "FCEEE3", [
        ("상세 견적", "'공사별로 품목·수량·단가까지 다 적힌 상세 견적서 주세요.'", "상세할수록 시공 실력이 좋고, 비교·가격 협상도 쉬워요", "필수"),
        ("현장 실측", "'현장 방문해서 실측하고 견적 주시는 거죠?'", "정확한 견적은 현장 실측 필수. 안 그러면 나중에 추가금이 커져요", "필수"),
        ("별도 항목", "'이 견적에 안 들어간 별도 항목이 뭐예요? (에어컨·가구·붙박이장·입주청소·승강기)'", "빠진 항목이 나중에 추가금이 됩니다. 4곳 다 가구·에어컨 별도 가능성 있음", "필수"),
        ("사기 주의", "댓글·오픈채팅방으로 결제 유도하면 절대 응하지 않기", "정식 업체는 댓글/오픈채팅 결제 요구를 안 합니다", "권장"),
    ]),
    ("2단계. 가격·결제", "F5DDA0", "FDF3E0", [
        ("부가세", "'이 금액 부가세 포함이에요, 별도예요?'", "리바트·한샘은 포함. 전화 3곳(아파트멘터리·퍼스트·홈루덴스)은 꼭 확인", "필수"),
        ("세금계산서", "'세금계산서(또는 현금영수증) 발행해 주세요.'", "부가세 합법 처리. 확장·샷시·단열·발코니확장은 양도세 필요경비 인정 / 싱크대·주방기구·도배·문·페인트는 불인정", "필수"),
        ("결제 분할", "'계약금-중도금-잔금으로 나눠 낼게요. 날짜·금액 계약서에 적어주세요.'", "큰돈 한 번에 주지 않기. 계약금은 공사비 20% 넘지 않게", "필수"),
        ("입금 계좌", "'입금은 사업자(법인/사장) 명의 계좌 맞죠?'", "담당자 개인계좌 입금은 위험. 현금 주면 영수증 즉시 받기", "권장"),
        ("추가비용", "'추가비용은 어떤 경우에 생겨요?'", "실측·자재 변경 시 변동. 조건을 계약 전에 알아두기", "필수"),
    ]),
    ("3단계. 계약·보증 (계약 전 확인)", "A9CCE3", "E8F1FB", [
        ("표준 계약서", "'표준 계약서 쓰고 견적서·상세내역·공정표 첨부해 주세요.'", "구두 약속은 무효. 전부 서면으로 남기기", "필수"),
        ("하자보증", "'하자보증 방법이랑 기간을 계약서에 적어주세요.'", "리바트 3년 / 퍼스트 1년 — 기간 반드시 명기", "필수"),
        ("보증증권", "'계약이행·하자이행 보증증권(서울보증) 발행되나요?'", "업체 부도·하자 불이행 대비. 하자보증금 5~10%가 적정", "권장"),
        ("지체보상금", "'공사 지연되면 지체보상금(1일당 공사비 1/1000) 넣어주세요.'", "우리집: 9월 입주(잔금 8/28). 공사 늦으면 이사 차질", "필수"),
        ("업체 신원", "'사업자등록증·신분증 확인할게요.'", "유령업체 방지. 계약서에 주소·연락처·담당자 명기", "권장"),
        ("확장 허가", "'발코니 확장 행위허가·관리실 동의 절차 진행되나요?'", "우리집: 확장 2개소. 구청 허가·관리사무소 신고 필요", "필수"),
    ]),
    ("4단계. 자재·사양 (우리집 공사)", "B6D7A8", "EAF4E6", [
        ("확장 단열", "'확장부 결로방지 단열 어떻게 하세요?'", "확장 2개소. 단열 안 하면 결로·곰팡이. 별도 단열 안 했으면 하자 인정 안 됨", "필수"),
        ("샷시 단열", "'샷시 교체하면서 주변 벽체 단열도 같이 하나요?'", "샷시 전체교체. 샷시만 좋으면 벽체에 결로가 더 생김 (단열 별도 공사)", "필수"),
        ("욕실 방수", "'욕실 방수 범위가 어디까지예요? 덧방이에요, 철거 후예요?'", "욕실 1개. 덧방=기존 방수 하자는 하자 아님 / 철거 후=방수 새로 함", "필수"),
        ("타일 시공", "'타일 덧방인지 철거 후 시공인지 알려주세요.'", "줄눈 2~6mm(베란다 6mm), 높이차 3mm까지는 허용오차", "권장"),
        ("도배", "'실크벽지는 띄움시공이라 가운데 들뜰 수 있죠? 기존벽지 제거 범위는요?'", "실크는 띄움시공이라 들뜸 가능. 미리 알아두기", "권장"),
        ("싱크대 높이", "'싱크대 상판 높이를 제 키(키÷2 +5cm)에 맞춰주세요.'", "설치 후 변경 불가. 2cm 오차 발생 가능", "권장"),
        ("목공 마감", "'몰딩·판재에 핀타카(타카핀) 자국 남나요?'", "마감 자국 여부 미리 확인", "참고"),
        ("에어컨", "'시스템에어컨 4대, 견적에 포함이에요 직접결제예요?'", "4대. 견적에 태우면 마진이 붙음(홈루덴스는 직접결제 권장)", "권장"),
    ]),
    ("5단계. 일정·공사 진행", "C9B6E4", "F1EAF9", [
        ("공정표", "'구체적인 공사 일정표(공정표) 주세요.'", "일정 관리·확인의 기준", "권장"),
        ("착공·준공", "'착공일·준공일 명확히, 예비공기 며칠 두고 잡아주세요.'", "9월 입주. 여유 공기를 둬야 이사에 안전", "필수"),
        ("직접 확인", "'설비·방수·단열·전기는 제가 직접 확인할게요.'", "가려지면 못 봅니다. 시공 중에 꼭 확인", "권장"),
        ("추가·변경 서면", "'공사 추가·변경되면 추가확인서(내용·비용·날짜·사인) 써주세요.'", "구두 변경이 분쟁 1순위", "필수"),
    ]),
    ("6단계. 완료·A/S", "A8DAD0", "E8F6F2", [
        ("완료확인서", "'공사완료확인서 쓰고 잔금 드릴게요. 완납확인서도 주세요.'", "완료 확인 후 잔금 지급", "권장"),
        ("A/S", "'하자 생기면 A/S 어떻게 받고, 기간은 얼마예요?'", "A/S 절차·기간을 계약 전에 명확히", "필수"),
    ]),
]

def imp_color6(s):
    return "BC8079" if s == "필수" else ("C39A4A" if s == "권장" else "9AA0A6")

r = hr6 + 1
for sec_title, hcolor, band, items in sections6:
    # 섹션 헤더 띠
    ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws6.cell(row=r, column=1, value=sec_title)
    c.fill = fill(hcolor); c.font = font(11, True, "33414F")
    c.alignment = Alignment(horizontal="left", vertical="center"); c.border = BORDER
    ws6.row_dimensions[r].height = 24
    r += 1
    for cat, q, why, imp in items:
        star = cat in {"확장 단열", "샷시 단열", "욕실 방수", "에어컨", "지체보상금", "확장 허가", "착공·준공"}
        label = ("[우리집] " + cat) if star else cat
        put(ws6, r, 1, label, fillc=("FBF0D8" if star else band), f=font(9, True, "9E7B4E" if star else "33414F"))
        put(ws6, r, 2, q, fillc="FFFFFF", f=font(10, True, "1F3A5F"))
        put(ws6, r, 3, why, fillc=band, f=font(9, color="555555"))
        put(ws6, r, 4, imp, fillc="FFFFFF", align=CENTER, f=font(10, True, imp_color6(imp)))
        put(ws6, r, 5, "", fillc="FFFFFF")
        ws6.row_dimensions[r].height = 36
        r += 1
ws6.freeze_panes = "A4"

# =====================================================================
# 시트 7 · 홈루덴스 방문상담 상세 (2026-06-13, 약 2시간)
# =====================================================================
ws7 = wb.create_sheet("7.홈루덴스_방문상담")
ws7.sheet_view.showGridLines = False
NC7 = 4
w7 = [16, 22, 54, 28]
for i, w in enumerate(w7, 1):
    ws7.column_dimensions[get_column_letter(i)].width = w

title_row(ws7, 1, NC7, "홈루덴스 방문상담 상세  ·  2026-06-13 (약 2시간)  ·  봉천동 두산 59A(24평형)", bg=SLATE)
note_row(ws7, 2, NC7,
         "전화 1차상담(6/11) 후 토요일 방문상담 녹음을 정리했습니다. 전문은 통화_전사/홈루덴스_방문상담록.md 참고. "
         "결정은 2차 미팅(6/18) 견적서에서 확정됩니다.", h=32)

hr7 = 3
for i, t in enumerate(["구분", "항목", "방문상담 결정 · 제안", "비고"], 1):
    put(ws7, hr7, i, t, fillc=HOME_H, f=font(11, True), align=CENTER)
ws7.row_dimensions[hr7].height = 26

rows7 = [
    ("집 진단", "현황", "순정(올수리 X) · 12층 계단식 · 갈매기몰딩 등 클래식 마감", "스프링클러 없음→단내림 목공 필요"),
    ("집 진단", "방향성", "무몰딩 · 30mm 얇은 걸레받이로 모던화, 골드/클래식 제거", "몰딩 철거 시 천장 처짐→목공 보강"),
    ("집 진단", "벽 상태", "수직·수평 불량. 공용부 한 면 면잡기 권장", "방은 가구로 가려 생략 가능"),
    ("현관", "중문", "12층 계단식이라 선택사항. 권장: 여닫이 3도어 (슬라이딩 불가-벽없음)", "넣기/빼기 견적 둘 다"),
    ("현관", "신발장", "띄움(플로팅) 신발장 + 하부 간접등. 오픈선반 X(수납 우선)", ""),
    ("주방", "후드", "위치 유지(배관). 빌트인 후드 권장(천장형보다 깔끔)", ""),
    ("주방", "가스/인덕션", "인덕션 사용 → 가스관을 베란다 계량기에서 차단(깔끔)", ""),
    ("주방", "냉장고", "키친핏(깊이700) vs 양문형(900). 공간상 키친핏 유리, 김냉 검토", "양문형 900은 부엌장 수납 죽음"),
    ("주방", "싱크 배치", "식기세척기600+인덕션600 폭 부족 가능성", "실측(CS) 때 확인"),
    ("주방", "식탁", "확장 시 6인/원형 가능. 사무 겸용 제안. 망장(양념서랍) 반영", ""),
    ("거실", "TV", "반매립은 고객 X 결정(공간 좁아짐). 콘센트 800높이로 정리", ""),
    ("거실", "에어컨/천장", "시스템 4대 + 단내림(간접등 처리). 커튼박스 양쪽 간접등", "전동커튼은 배선만 시공"),
    ("방", "안방", "침대 퀸. 시스템장(400)+행거 복합. 수납 일단 제외(혼자 선거주)", "장 짜기/빼기 견적 둘 다"),
    ("방", "기타방·베란다", "컴퓨터방·운동방 배치. 앞방 베란다=창고+도장+빨래건조", "외부 빨래 불가→실외기실 문짝(관리실 확인)"),
    ("욕실(1개)", "구성", "롱젠다이 · 졸리컷 타일(비드X) · 유리파티션 투명+밑띄움 · 욕조→샤워", ""),
    ("욕실(1개)", "거울장/전기", "슬라이딩 거울장(800) · 내부 간접등+콘센트 · 비데콘센트 정리", ""),
    ("욕실(1개)", "환기건조", "휴젠뜨/제로크/하츠 vs 일반+전동댐퍼. 고객: 일단 일반+전동댐퍼", "5배 차이(휴젠뜨30~40 vs 7~8만)·추후 조정"),
    ("마감재", "벽지", "개나리 로하스/프리모 or LX 디아망/베스띠 or 신한. 회벽톤 유행→개나리 방향", "웜/쿨/그레이지 정하면 선택 축소"),
    ("마감재", "마루", "와이드 강마루(동화 나투스진그란데/한솔SB ~11만/평) vs LX 에디터온 SPC(찍힘 강)", "컬러 우선 선택"),
    ("마감재", "창호", "HOMECC(홈씨씨) 확정 — 가스켓 방식, 곰팡이↓", "LX 뷰프레임/KCC 대비 중간가"),
    ("마감재", "가구", "한샘(부엌·현관장). 욕실장은 제작(방수판)", ""),
    ("견적·일정", "예산", "고객 상한 6천 초반(VAT포함). 홈루덴스: 그 안에서 맞춤", "타사: 리바트5,525 / 한샘7,812 / 5천대·7천 각1곳"),
    ("견적·일정", "에어컨", "견적 제외 → 직접결제(마진절감). 단내림 목공은 포함", ""),
    ("견적·일정", "2차 미팅", "6/18(목) 19시, 서초(남부터미널) 비상주 사무실. 3D 설계 제시", ""),
    ("견적·일정", "실측", "픽스 후 창호팀+소장 동반. 거주 어르신 양해 필요", "16일 리바트 실측 예정"),
    ("견적·일정", "공사", "8월말 착공 · 잔금 8/28 · 입주 9/30", ""),
    ("⭐디자이너 조언", "업체검증", "건설면허 보유 확인(없으면 1,500만 한도)", "업체선정 일반 팁"),
    ("⭐디자이너 조언", "견적비교", "너무 싼 1곳 빼고 나머지 평균→수렴값. 견적서 맨밑 잡비·마진(25%) 확인", ""),
    ("⭐디자이너 조언", "추가금", "철거 후 A/B/C안 중 C(안함)가 베스트. 접지없음·전선1.5sq 등 특이사항 비용 가능", ""),
    ("⭐디자이너 조언", "직접구매", "에어컨·인터폰·커튼·러그=공산품 최저가 직접(러그는 워셔블)", ""),
    ("⭐디자이너 조언", "디자인", "컬러/패턴 세게 X(2~3년 질림). 포인트는 쿠션·러그·소품으로", ""),
]

band7 = {"집 진단": "FBF7EF", "현관": "F0F5FA", "주방": "F1F8EE", "거실": "FAF1F7",
         "방": "F3F0FA", "욕실(1개)": "FDF3E9", "마감재": "EAF4F0", "견적·일정": "FCEEE3"}
r = hr7 + 1
for cat, item, body, memo in rows7:
    star = cat.startswith("⭐")
    cbg = "FBF0D8" if star else band7.get(cat, "FFFFFF")
    put(ws7, r, 1, cat, fillc=("FBF0D8" if star else CREAM), f=font(9, True, "9E7B4E" if star else "2B2B2B"))
    put(ws7, r, 2, item, fillc=cbg, f=font(10, True))
    put(ws7, r, 3, body, fillc="FFFFFF", f=font(10))
    put(ws7, r, 4, memo, fillc=cbg, f=font(9, color="666666"))
    ws7.row_dimensions[r].height = 30
    r += 1
ws7.freeze_panes = "C4"

import os
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "인테리어_업체비교.xlsx")
wb.save(OUT)
print("저장 완료:", OUT)
