from pathlib import Path

from openpyxl import load_workbook

from scripts.content_attachments.excel import write_excel


def test_write_excel_produces_file(tmp_path: Path):
    dataset = {
        "title": "이재명 대통령 당선후 서울 실거래 변화",
        "periodLabel": "2024.6 ~ 2025.6 vs 2025.6 ~ 현재",
        "source": "국토부 실거래가 (매매)",
        "districts": [
            {"district": "광진구", "priceBefore": 14487, "priceAfter": 16997, "changePct": 17.3},
            {"district": "강남구", "priceBefore": 26318, "priceAfter": 26915, "changePct": 2.3},
        ],
    }
    out = tmp_path / "out.xlsx"
    write_excel(dataset, out)
    assert out.exists()

    wb = load_workbook(out)
    ws = wb.active
    assert ws["A1"].value == "이재명 대통령 당선후 서울 실거래 변화"
    assert ws["A3"].value == "지역"
    assert ws["B3"].value == "취임 전 (만원)"
    assert ws["C3"].value == "취임 후 (만원)"
    assert ws["D3"].value == "변동률(%)"
    assert ws["A4"].value == "광진구"
    assert ws["B4"].value == 14487
    assert ws["C4"].value == 16997
    assert ws["D4"].value == 17.3


def test_write_excel_25_districts(tmp_path: Path):
    dataset = {
        "title": "t",
        "periodLabel": "",
        "source": "",
        "districts": [
            {"district": f"{i}구", "priceBefore": 100, "priceAfter": 110, "changePct": 10.0}
            for i in range(25)
        ],
    }
    out = tmp_path / "out.xlsx"
    write_excel(dataset, out)
    wb = load_workbook(out)
    ws = wb.active
    # Last data row at A28 (3 header rows + 25 districts)
    assert ws["A28"].value == "24구"
